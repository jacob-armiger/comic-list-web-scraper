#! python3
import requests, bs4
from flask import flash
from openpyxl import Workbook
import urllib3

def scrape(url):
    """
    Takes a URL and scrapes it for <p> elements
    Returns bs4 object
    """
    http = urllib3.PoolManager()
    try:
        page = http.request("GET",url)
    except requests.ConnectionError as e:
        flash("We couldn't connect to that URL. Try again!\n")
        print(str(e))            
        return -1
    except requests.Timeout as e:
        flash("OOPS!! Timeout Error")
        print(str(e))
        return -1
    except requests.RequestException as e:
        flash("OOPS!! General Error")
        print(str(e))
        return -1

    # Parse HTML to find all <p> elements
    soup = bs4.BeautifulSoup(page.data, 'lxml')
    soup.prettify()
    soup_list_object = soup.find_all('p')

    return soup_list_object

def create_csv(soup_list_obj): 
    """
    Takes list of <p> elements and separates them into comic titles.
    Then joins titles into a CSV file.
    Returns CSV file name
    """

    FILE_NAME = "reading_list.txt"
    comic_list = []

    # Loops through every <p> element
    for p_element in soup_list_obj:
        # This splits <p> elements that have multiple comic titles
        chunk = p_element.text.split('\n')

        for comic in chunk:
            # Add comic book event listings to list
            if(comic.find(" here.") != -1):
                comic_list.append(comic.lstrip() + ";")
                continue
            # Do not add
            if(comic.find("First Appearance:") != -1):
                continue
            # Add alternate starts
            if(comic.find("Alternate Starting Point:") != -1):
                comic_list.append(comic.lstrip() + ";")
                continue
            # Add comics that are numbered or dated
            for char in comic:
                if((char == '#') or (char == '(')):
                    comic_list.append(comic.lstrip() + ";")
                    break

    
    # Re-join comic titles into a CSV file
    csv_formatted_string = ''.join(comic_list)

    # Opens file to write CSV formatted data
    file = open("reading_list.txt", 'w')
    file.write(csv_formatted_string)

    file.close()

    return FILE_NAME


def create_excel(csv_file_name):
    """
    Create excel worksheet from a CSV file.
    Returns excel file name
    """
    wb = Workbook()
    ws = wb.create_sheet("Comics",0)

    # Read CSV file
    file = open(csv_file_name, 'r')
    text = file.read()

    # Create array of comics titles with semi-colon delimiter
    comics = text.split(';')

    # Get max length of comic titles
    max_length = max(comics, key=len)

    # Assign values to cells
    for index, comic in enumerate(comics):
        ws.cell(row=index+1, column=1, value=comic)

    # Set proper width for column
    ws.column_dimensions['A'].width = len(max_length)
    EXCEL_FILE_NAME = "reading_list.xlsx"

    wb.save(EXCEL_FILE_NAME)

    return EXCEL_FILE_NAME


"""
def main():
    # URL to be scraped
    REQUEST_URL = ''

    # Get input from the user
    print("Welcome! Copy and paste a reading order from \nhttps://comicbookreadingorders.com/ or click enter\n")
    REQUEST_URL = input("URL: ");

    # Use URL from the user or a default URL
    if REQUEST_URL:
        pass
    elif not REQUEST_URL:
        REQUEST_URL = 'https://comicbookreadingorders.com/dc/event-timeline/'


    # Stores comic book titles listed on URL
    text = scrape(REQUEST_URL)


    create_csv(text)
    create_excel("reading_list.txt")


if __name__ == "__main__":
    main()
"""





